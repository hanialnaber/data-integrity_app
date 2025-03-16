import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
import streamlit as st

# Define lighter colors for highlighting
RED_FILL = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")  # Light red for Missing
YELLOW_FILL = PatternFill(start_color="FFFFEB99", end_color="FFFFEB99", fill_type="solid")  # Light yellow for Value difference
GREEN_FILL = PatternFill(start_color="FF99FF99", end_color="FF99FF99", fill_type="solid")  # Light green for Extra
BLUE_FILL = PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid")  # Light blue for Order mismatch

def highlight_differences_excel(file1_path, file2_path, output_path, error_details):
    """
    Create a highlighted Excel file showing differences between two Excel files.
    
    Args:
        file1_path: Path to first Excel file
        file2_path: Path to second Excel file
        output_path: Path to save highlighted Excel file
        error_details: Dictionary with details of errors
        
    Returns:
        Path to highlighted Excel file
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment
    import streamlit as st
    
    # Create color fills
    RED_FILL = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")  # Light red
    YELLOW_FILL = PatternFill(start_color="FFFFEB99", end_color="FFFFEB99", fill_type="solid")  # Light yellow
    GREEN_FILL = PatternFill(start_color="FF99FF99", end_color="FF99FF99", fill_type="solid")  # Light green
    BLUE_FILL = PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid")  # Light blue
    
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Read the first Excel file
    try:
        data1 = pd.read_excel(file1_path, sheet_name=None)
    except Exception as e:
        st.error(f"Error reading first Excel file: {str(e)}")
        return None
    
    # Read the second Excel file
    try:
        data2 = pd.read_excel(file2_path, sheet_name=None)
    except Exception as e:
        st.error(f"Error reading second Excel file: {str(e)}")
        return None
    
    # Get all sheet names
    all_sheets = set(list(data1.keys()) + list(data2.keys()))
    
    # Process each sheet
    for sheet in all_sheets:
        sheet_str = str(sheet)  # Convert to string to ensure it's hashable
        
        # Create a new sheet
        worksheet = wb.create_sheet(title=sheet_str[:31])  # Excel sheet names are limited to 31 chars
        
        # Check if sheet exists in both files
        if sheet_str in data1 and sheet_str in data2:
            df1 = data1[sheet_str]
            df2 = data2[sheet_str]
            
            # Get all columns
            all_cols = list(set(list(df1.columns) + list(df2.columns)))
            
            # Write header row
            for i, col in enumerate(all_cols):
                cell = worksheet.cell(row=1, column=i+1)
                cell.value = str(col)
                cell.font = Font(bold=True)
            
            # Create a mapping of column names to indices
            col_indices = {col: i+1 for i, col in enumerate(all_cols)}
            
            # Write data from first file
            for i, row in df1.iterrows():
                for col in df1.columns:
                    if col in col_indices:
                        cell = worksheet.cell(row=i+2, column=col_indices[col])
                        cell.value = str(row[col])
            
            # Highlight missing sheets
            if "missing_sheets" in error_details and sheet_str in error_details["missing_sheets"]:
                # Highlight entire sheet
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.fill = RED_FILL
            
            # Highlight extra sheets
            if "extra_sheets" in error_details and sheet_str in error_details["extra_sheets"]:
                # Highlight entire sheet
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.fill = GREEN_FILL
            
            # Highlight column differences
            if "column_differences" in error_details and sheet_str in error_details["column_differences"]:
                col_diffs = error_details["column_differences"][sheet_str]
                
                # Highlight missing columns
                if "missing_columns" in col_diffs:
                    for col in col_diffs["missing_columns"]:
                        if col in col_indices:
                            for row in range(1, worksheet.max_row + 1):
                                cell = worksheet.cell(row=row, column=col_indices[col])
                                cell.fill = RED_FILL
                
                # Highlight extra columns
                if "extra_columns" in col_diffs:
                    for col in col_diffs["extra_columns"]:
                        if col in col_indices:
                            for row in range(1, worksheet.max_row + 1):
                                cell = worksheet.cell(row=row, column=col_indices[col])
                                cell.fill = GREEN_FILL
                
                # Highlight reordered columns
                if "reordered_columns" in col_diffs and col_diffs["reordered_columns"]:
                    for col in col_diffs["reordered_columns"]:
                        if col in col_indices:
                            cell = worksheet.cell(row=1, column=col_indices[col])
                            cell.fill = BLUE_FILL
            
            # Highlight row differences
            if "row_differences" in error_details and sheet_str in error_details["row_differences"]:
                row_diffs = error_details["row_differences"][sheet_str]
                
                # Highlight missing rows
                if "missing_rows" in row_diffs:
                    missing_rows = row_diffs["missing_rows"]
                    if isinstance(missing_rows, list):
                        # Handle list of dictionaries (new format)
                        for row_info in missing_rows:
                            # Try to find the row in the dataframe
                            found = False
                            for i, row in df1.iterrows():
                                match = True
                                for k, v in row_info.items():
                                    if k in row and str(row[k]) != str(v):
                                        match = False
                                        break
                                if match:
                                    # Highlight the row
                                    for col in df1.columns:
                                        if col in col_indices:
                                            cell = worksheet.cell(row=i+2, column=col_indices[col])
                                            cell.fill = RED_FILL
                                    found = True
                                    break
                    elif isinstance(missing_rows, dict):
                        # Handle dictionary format (old format)
                        for key, row_idx in missing_rows.items():
                            # Highlight the row
                            for col in df1.columns:
                                if col in col_indices:
                                    cell = worksheet.cell(row=row_idx+2, column=col_indices[col])
                                    cell.fill = RED_FILL
                
                # Highlight extra rows
                if "extra_rows" in row_diffs:
                    extra_rows = row_diffs["extra_rows"]
                    if isinstance(extra_rows, list):
                        # Handle list of dictionaries (new format)
                        for row_info in extra_rows:
                            # Try to find the row in the dataframe
                            found = False
                            for i, row in df2.iterrows():
                                match = True
                                for k, v in row_info.items():
                                    if k in row and str(row[k]) != str(v):
                                        match = False
                                        break
                                if match:
                                    # Highlight the row
                                    row_idx = i + df1.shape[0] + 2
                                    for col in df2.columns:
                                        if col in col_indices:
                                            cell = worksheet.cell(row=row_idx, column=col_indices[col])
                                            cell.value = str(row[col])
                                            cell.fill = GREEN_FILL
                                    found = True
                                    break
                    elif isinstance(extra_rows, dict):
                        # Handle dictionary format (old format)
                        for key, row_idx in extra_rows.items():
                            # Get the row from the second dataframe
                            if row_idx < len(df2):
                                row = df2.iloc[row_idx]
                                # Add the row to the worksheet
                                row_idx = row_idx + df1.shape[0] + 2
                                for col in df2.columns:
                                    if col in col_indices:
                                        cell = worksheet.cell(row=row_idx, column=col_indices[col])
                                        cell.value = str(row[col])
                                        cell.fill = GREEN_FILL
            
                # Highlight value differences
                if sheet_str in error_details["value_differences"]:
                    value_diffs = error_details["value_differences"][sheet_str]
                    
                    # Limit the number of differences to highlight (for performance)
                    if len(value_diffs) > 1000:
                        value_diffs = value_diffs[:1000]
                    
                    # Create a lookup dictionary for faster access
                    diff_lookup = {}
                    
                    # Process each difference
                    for diff in value_diffs:
                        # Get the key (row or key value)
                        if isinstance(diff.get("key"), dict):
                            # If key is a dictionary, convert it to a string representation
                            key_parts = []
                            for k, v in diff.get("key").items():
                                key_parts.append(f"{k}={v}")
                            key = "|".join(key_parts)
                        else:
                            key = diff.get("key", diff.get("row"))
                        
                        # Get the column
                        col = diff.get("column")
                        
                        # Skip if column not found
                        if col not in col_indices:
                            continue
                        
                        # Add to lookup dictionary - ensure key is hashable
                        str_key = str(key)
                        if str_key not in diff_lookup:
                            diff_lookup[str_key] = {}
                        
                        diff_lookup[str_key][col] = {
                            "value1": diff.get("value1"),
                            "value2": diff.get("value2")
                        }
                    
                    # Process key-based differences
                    key_based_count = 0
                    if any(isinstance(diff.get("key"), dict) for diff in value_diffs if "key" in diff):
                        # Get the key column(s)
                        # Try to extract key column names from the first difference with a key
                        key_cols = []
                        for diff in value_diffs:
                            if isinstance(diff.get("key"), dict):
                                key_cols = list(diff.get("key").keys())
                                break
                        
                        if not key_cols:
                            key_cols = [df1.columns[0]]  # Fallback to first column
                        
                        # Create a mapping from key to row index
                        key_to_row = {}
                        
                        # For each row in the dataframe
                        for i, row in df1.iterrows():
                            # Create a key string similar to how we created it above
                            key_parts = []
                            for col in key_cols:
                                if col in row:
                                    key_parts.append(f"{col}={row[col]}")
                            key_str = "|".join(key_parts)
                            key_to_row[key_str] = i
                        
                        # Process differences
                        for key, cols in diff_lookup.items():
                            if key in key_to_row:  # It's a key-based diff
                                row_idx = key_to_row[key] + 2  # +2 for header and 1-indexing
                                
                                for col_name, diff in cols.items():
                                    if col_name in col_indices:
                                        col_idx = col_indices[col_name]
                                        
                                        # Highlight the cell
                                        cell = worksheet.cell(row=row_idx, column=col_idx)
                                        cell.fill = YELLOW_FILL
                                        
                                        # Add a comment with the difference
                                        comment_text = f"Value in file 1: {diff['value1']}\nValue in file 2: {diff['value2']}"
                                        cell.comment = Comment(comment_text, "Comparison Ability")
                                        
                                        key_based_count += 1
                    
                    # Process row-based differences
                    row_based_count = 0
                    for key, cols in diff_lookup.items():
                        try:
                            # Check if key can be converted to integer (row index)
                            row = int(key)
                            row_idx = row + 2  # +2 for header and 1-indexing
                            
                            for col_name, diff in cols.items():
                                if col_name in col_indices:
                                    col_idx = col_indices[col_name]
                                    
                                    # Highlight the cell
                                    cell = worksheet.cell(row=row_idx, column=col_idx)
                                    cell.fill = YELLOW_FILL
                                    
                                    # Add a comment with the difference
                                    comment_text = f"Value in file 1: {diff['value1']}\nValue in file 2: {diff['value2']}"
                                    cell.comment = Comment(comment_text, "Comparison Ability")
                                    
                                    row_based_count += 1
                        except (ValueError, TypeError):
                            # Not a row-based diff, skip
                            continue
        else:
            # Sheet only exists in one file
            if sheet_str in data1:
                df = data1[sheet_str]
                
                # Write header row
                for i, col in enumerate(df.columns):
                    cell = worksheet.cell(row=1, column=i+1)
                    cell.value = str(col)
                    cell.font = Font(bold=True)
                
                # Write data
                for i, row in df.iterrows():
                    for j, col in enumerate(df.columns):
                        cell = worksheet.cell(row=i+2, column=j+1)
                        cell.value = str(row[col])
                        cell.fill = RED_FILL  # Missing in second file
            else:
                df = data2[sheet_str]
                
                # Write header row
                for i, col in enumerate(df.columns):
                    cell = worksheet.cell(row=1, column=i+1)
                    cell.value = str(col)
                    cell.font = Font(bold=True)
                
                # Write data
                for i, row in df.iterrows():
                    for j, col in enumerate(df.columns):
                        cell = worksheet.cell(row=i+2, column=j+1)
                        cell.value = str(row[col])
                        cell.fill = GREEN_FILL  # Extra in second file
    
    # Add a summary sheet
    summary = wb.create_sheet(title="Summary", index=0)
    
    # Add color legend
    summary.cell(row=1, column=1).value = "Color Legend"
    summary.cell(row=1, column=1).font = Font(bold=True)
    
    summary.cell(row=2, column=1).value = "Missing in second file"
    summary.cell(row=2, column=1).fill = RED_FILL
    
    summary.cell(row=3, column=1).value = "Value differences"
    summary.cell(row=3, column=1).fill = YELLOW_FILL
    
    summary.cell(row=4, column=1).value = "Extra in second file"
    summary.cell(row=4, column=1).fill = GREEN_FILL
    
    summary.cell(row=5, column=1).value = "Order mismatch"
    summary.cell(row=5, column=1).fill = BLUE_FILL
    
    # Save the workbook
    try:
        wb.save(output_path)
        return output_path
    except Exception as e:
        st.error(f"Error saving highlighted Excel file: {str(e)}")
        return None

def highlight_differences_csv(file1_path, file2_path, output_path, error_details):
    """
    Create a highlighted Excel file showing differences between two CSV files.
    
    Args:
        file1_path: Path to first CSV file
        file2_path: Path to second CSV file
        output_path: Path to save highlighted Excel file
        error_details: Dictionary with details of errors
        
    Returns:
        Path to highlighted Excel file
    """
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment
    import streamlit as st
    
    # Create color fills
    RED_FILL = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")  # Light red
    YELLOW_FILL = PatternFill(start_color="FFFFEB99", end_color="FFFFEB99", fill_type="solid")  # Light yellow
    GREEN_FILL = PatternFill(start_color="FF99FF99", end_color="FF99FF99", fill_type="solid")  # Light green
    BLUE_FILL = PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid")  # Light blue
    
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Read the first CSV file
    try:
        df1 = pd.read_csv(file1_path)
    except Exception as e:
        st.error(f"Error reading first CSV file: {str(e)}")
        return None
    
    # Read the second CSV file
    try:
        df2 = pd.read_csv(file2_path)
    except Exception as e:
        st.error(f"Error reading second CSV file: {str(e)}")
        return None
    
    # Create a new sheet
    worksheet = wb.create_sheet(title="Data")
    
    # Get all columns
    all_cols = list(set(list(df1.columns) + list(df2.columns)))
    
    # Write header row
    for i, col in enumerate(all_cols):
        cell = worksheet.cell(row=1, column=i+1)
        cell.value = str(col)
        cell.font = Font(bold=True)
    
    # Create a mapping of column names to indices
    col_indices = {col: i+1 for i, col in enumerate(all_cols)}
    
    # Write data from first file
    for i, row in df1.iterrows():
        for col in df1.columns:
            if col in col_indices:
                cell = worksheet.cell(row=i+2, column=col_indices[col])
                cell.value = str(row[col])
    
    # Highlight column differences
    if "column_differences" in error_details and "data" in error_details["column_differences"]:
        col_diffs = error_details["column_differences"]["data"]
        
        # Highlight missing columns
        if "missing_columns" in col_diffs:
            for col in col_diffs["missing_columns"]:
                if col in col_indices:
                    for row in range(1, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_indices[col])
                        cell.fill = RED_FILL
        
        # Highlight extra columns
        if "extra_columns" in col_diffs:
            for col in col_diffs["extra_columns"]:
                if col in col_indices:
                    for row in range(1, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_indices[col])
                        cell.fill = GREEN_FILL

        # Highlight reordered columns
        if "reordered_columns" in col_diffs and col_diffs["reordered_columns"]:
            for col in col_diffs["reordered_columns"]:
                if col in col_indices:
                    cell = worksheet.cell(row=1, column=col_indices[col])
                    cell.fill = BLUE_FILL

    # Highlight row differences
    if "row_differences" in error_details and "data" in error_details["row_differences"]:
        row_diffs = error_details["row_differences"]["data"]
        
        # Highlight missing rows
        if "missing_rows" in row_diffs:
            missing_rows = row_diffs["missing_rows"]
            if isinstance(missing_rows, list):
                # Handle list of dictionaries (new format)
                for row_info in missing_rows:
                    # Try to find the row in the dataframe
                    found = False
                    for i, row in df1.iterrows():
                        match = True
                        for k, v in row_info.items():
                            if k in row and str(row[k]) != str(v):
                                match = False
                                break
                        if match:
                            # Highlight the row
                            for col in df1.columns:
                                if col in col_indices:
                                    cell = worksheet.cell(row=i+2, column=col_indices[col])
                                    cell.fill = RED_FILL
                            found = True
                            break
            elif isinstance(missing_rows, dict):
                # Handle dictionary format (old format)
                for key, row_idx in missing_rows.items():
                    # Highlight the row
                    for col in df1.columns:
                        if col in col_indices:
                            cell = worksheet.cell(row=row_idx+2, column=col_indices[col])
                            cell.fill = RED_FILL
        
        # Highlight extra rows
        if "extra_rows" in row_diffs:
            extra_rows = row_diffs["extra_rows"]
            if isinstance(extra_rows, list):
                # Handle list of dictionaries (new format)
                for row_info in extra_rows:
                    # Try to find the row in the dataframe
                    found = False
                    for i, row in df2.iterrows():
                        match = True
                        for k, v in row_info.items():
                            if k in row and str(row[k]) != str(v):
                                match = False
                                break
                        if match:
                            # Highlight the row
                            row_idx = i + df1.shape[0] + 2
                            for col in df2.columns:
                                if col in col_indices:
                                    cell = worksheet.cell(row=row_idx, column=col_indices[col])
                                    cell.value = str(row[col])
                                    cell.fill = GREEN_FILL
                            found = True
                            break
            elif isinstance(extra_rows, dict):
                # Handle dictionary format (old format)
                for key, row_idx in extra_rows.items():
                    # Get the row from the second dataframe
                    if row_idx < len(df2):
                        row = df2.iloc[row_idx]
                        # Add the row to the worksheet
                        row_idx = row_idx + df1.shape[0] + 2
                        for col in df2.columns:
                            if col in col_indices:
                                cell = worksheet.cell(row=row_idx, column=col_indices[col])
                                cell.value = str(row[col])
                                cell.fill = GREEN_FILL
    
    # Highlight value differences
    if "value_differences" in error_details and "data" in error_details["value_differences"]:
        value_diffs = error_details["value_differences"]["data"]
        
        # Limit the number of differences to highlight (for performance)
        if len(value_diffs) > 1000:
            value_diffs = value_diffs[:1000]
        
        # Create a lookup dictionary for faster access
        diff_lookup = {}
        
        # Process each difference
        for diff in value_diffs:
            # Get the key (row or key value)
            if isinstance(diff.get("key"), dict):
                # If key is a dictionary, convert it to a string representation
                key_parts = []
                for k, v in diff.get("key").items():
                    key_parts.append(f"{k}={v}")
                key = "|".join(key_parts)
            else:
                key = diff.get("key", diff.get("row"))
            
            # Get the column
            col = diff.get("column")
            
            # Skip if column not found
            if col not in col_indices:
                continue
            
            # Add to lookup dictionary - ensure key is hashable
            str_key = str(key)
            if str_key not in diff_lookup:
                diff_lookup[str_key] = {}
            
            diff_lookup[str_key][col] = {
                "value1": diff.get("value1"),
                "value2": diff.get("value2")
            }
        
        # Process key-based differences
        key_based_count = 0
        if any(isinstance(diff.get("key"), dict) for diff in value_diffs if "key" in diff):
            # Get the key column(s)
            # Try to extract key column names from the first difference with a key
            key_cols = []
            for diff in value_diffs:
                if isinstance(diff.get("key"), dict):
                    key_cols = list(diff.get("key").keys())
                    break
            
            if not key_cols:
                key_cols = [df1.columns[0]]  # Fallback to first column
            
            # Create a mapping from key to row index
            key_to_row = {}
            
            # For each row in the dataframe
            for i, row in df1.iterrows():
                # Create a key string similar to how we created it above
                key_parts = []
                for col in key_cols:
                    if col in row:
                        key_parts.append(f"{col}={row[col]}")
                key_str = "|".join(key_parts)
                key_to_row[key_str] = i
            
            # Process differences
            for key, cols in diff_lookup.items():
                if key in key_to_row:  # It's a key-based diff
                    row_idx = key_to_row[key] + 2  # +2 for header and 1-indexing
                    
                    for col_name, diff in cols.items():
                        if col_name in col_indices:
                            col_idx = col_indices[col_name]
                            
                            # Highlight the cell
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = YELLOW_FILL
                            
                            # Add a comment with the difference
                            comment_text = f"Value in file 1: {diff['value1']}\nValue in file 2: {diff['value2']}"
                            cell.comment = Comment(comment_text, "Comparison Ability")
                            
                            key_based_count += 1
        
        # Process row-based differences
        row_based_count = 0
        for key, cols in diff_lookup.items():
            try:
                # Check if key can be converted to integer (row index)
                row = int(key)
                row_idx = row + 2  # +2 for header and 1-indexing
                
                for col_name, diff in cols.items():
                    if col_name in col_indices:
                        col_idx = col_indices[col_name]
                        
                        # Highlight the cell
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.fill = YELLOW_FILL
                        
                        # Add a comment with the difference
                        comment_text = f"Value in file 1: {diff['value1']}\nValue in file 2: {diff['value2']}"
                        cell.comment = Comment(comment_text, "Comparison Ability")
                        
                        row_based_count += 1
            except (ValueError, TypeError):
                # Not a row-based diff, skip
                continue
    
    # Add a summary sheet
    summary = wb.create_sheet(title="Summary", index=0)
    
    # Add color legend
    summary.cell(row=1, column=1).value = "Color Legend"
    summary.cell(row=1, column=1).font = Font(bold=True)
    
    summary.cell(row=2, column=1).value = "Missing in second file"
    summary.cell(row=2, column=1).fill = RED_FILL
    
    summary.cell(row=3, column=1).value = "Value differences"
    summary.cell(row=3, column=1).fill = YELLOW_FILL
    
    summary.cell(row=4, column=1).value = "Extra in second file"
    summary.cell(row=4, column=1).fill = GREEN_FILL
    
    summary.cell(row=5, column=1).value = "Order mismatch"
    summary.cell(row=5, column=1).fill = BLUE_FILL
    
    # Save the workbook
    try:
        wb.save(output_path)
        return output_path
    except Exception as e:
        st.error(f"Error saving highlighted Excel file: {str(e)}")
        return None